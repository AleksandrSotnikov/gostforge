"""CLI-интерфейс gostforge."""

from __future__ import annotations

import json
import os
import subprocess
import sys
import time
from collections import defaultdict
from pathlib import Path
from typing import Any

import click

from gostforge import __version__
from gostforge.exporter import export_docx
from gostforge.fixer import FixApplied
from gostforge.fixer import fix as run_fix
from gostforge.parser import parse_docx
from gostforge.profile import list_profiles, load_profile
from gostforge.validator import Violation, validate
from gostforge.validator.engine import registered_checks

# Цветовая палитра по серьёзности нарушения. Click автоматически отключает
# цвет, если stdout не терминал, так что отдельно проверять не нужно.
_SEVERITY_STYLE = {
    "error": ("ошибок", {"fg": "red", "bold": True}, "ERROR"),
    "warning": ("предупр.", {"fg": "yellow", "bold": True}, "WARN "),
    "info": ("инфо", {"fg": "cyan"}, "INFO "),
}


def _print_violations(target: Path, violations: list[Violation], quiet: bool) -> None:
    """Сгруппировать и красиво вывести violations для одного документа."""
    click.secho(f"\n>>> {target.name}", bold=True)

    if not violations:
        click.echo("  " + click.style("[OK]", fg="green", bold=True) + "  Нарушений не найдено")
        return

    by_severity: dict[str, list[Violation]] = defaultdict(list)
    for v in violations:
        by_severity[v.severity].append(v)

    counts = {sev: len(items) for sev, items in by_severity.items()}
    summary = ", ".join(
        f"{counts.get(sev, 0)} {label}"
        for sev, (label, _style, _short) in _SEVERITY_STYLE.items()
        if counts.get(sev, 0) > 0
    )
    click.echo(
        "  " + click.style("[FAIL]", fg="red", bold=True) + f"  Найдено нарушений: {len(violations)} ({summary})"
    )

    if quiet:
        codes = sorted({v.check_code for v in violations})
        click.echo("  Коды: " + ", ".join(codes))
        return

    for severity in ("error", "warning", "info"):
        items = by_severity.get(severity, [])
        if not items:
            continue
        _, style, short = _SEVERITY_STYLE[severity]
        for v in items:
            tag = click.style(f"[{short}]", **style)
            code = click.style(f"{v.check_code}", bold=True)
            click.echo(f"  {tag} {code}  {v.message}")
            if v.location:
                click.echo("       " + click.style(v.location, fg="bright_black"))
            if v.suggestion:
                click.echo("       " + click.style("-> " + v.suggestion, fg="green"))


def _write_markdown_report(
    results: dict[str, list[Violation]], output: Path, profile_id: str
) -> None:
    """Сохранить отчёт в Markdown с группировкой по файлам и категориям проверок."""
    lines: list[str] = []
    lines.append(f"# Отчёт нормоконтроля — профиль `{profile_id}`")
    lines.append("")

    total = sum(len(v) for v in results.values())
    files_with_violations = sum(1 for v in results.values() if v)
    lines.append(
        f"Проверено файлов: **{len(results)}**, с нарушениями: **{files_with_violations}**, "
        f"всего нарушений: **{total}**"
    )
    lines.append("")

    for file_path, violations in results.items():
        name = Path(file_path).name
        lines.append(f"## {name}")
        lines.append("")
        if not violations:
            lines.append("Нарушений не найдено.")
            lines.append("")
            continue

        # Группируем по префиксу кода (категория)
        by_category: dict[str, list[Violation]] = defaultdict(list)
        for v in violations:
            category = v.check_code.split(".")[0]
            by_category[category].append(v)

        for category in sorted(by_category):
            lines.append(f"### Категория {category}")
            lines.append("")
            lines.append("| Код | Серьёзность | Сообщение | Что исправить |")
            lines.append("| --- | --- | --- | --- |")
            for v in by_category[category]:
                msg = v.message.replace("|", "\\|")
                sug = v.suggestion.replace("|", "\\|") if v.suggestion else ""
                lines.append(f"| {v.check_code} | {v.severity} | {msg} | {sug} |")
            lines.append("")

    output.write_text("\n".join(lines) + "\n", encoding="utf-8")


def _write_xlsx_report(
    results: dict[str, list[Violation]], output: Path, profile_id: str
) -> None:
    """Сохранить отчёт в Excel: один лист «Сводка» + по листу на каждый файл."""
    # openpyxl уже в зависимостях; импорт локально, чтобы CLI грузился без него
    # при использовании только Markdown-отчётов.
    from openpyxl import Workbook  # type: ignore[import-not-found]
    from openpyxl.styles import Alignment, Font, PatternFill  # type: ignore[import-not-found]

    wb = Workbook()
    summary = wb.active
    summary.title = "Сводка"

    bold = Font(bold=True)
    header_fill = PatternFill(start_color="DDDDDD", end_color="DDDDDD", fill_type="solid")
    wrap = Alignment(wrap_text=True, vertical="top")

    summary.append([f"Отчёт нормоконтроля — профиль {profile_id}"])
    summary["A1"].font = Font(bold=True, size=14)
    summary.append([])
    summary.append(["Файл", "Нарушений", "Ошибок", "Предупр.", "Инфо"])
    for cell in summary[3]:
        cell.font = bold
        cell.fill = header_fill

    for file_path, violations in results.items():
        name = Path(file_path).name
        total = len(violations)
        errs = sum(1 for v in violations if v.severity == "error")
        warns = sum(1 for v in violations if v.severity == "warning")
        infos = sum(1 for v in violations if v.severity == "info")
        summary.append([name, total, errs, warns, infos])

    for col_letter, width in zip("ABCDE", (40, 12, 10, 12, 8), strict=True):
        summary.column_dimensions[col_letter].width = width

    for file_path, violations in results.items():
        # Имя листа — обрезаем до 31 символа (ограничение Excel) и удаляем спецсимволы
        sheet_name = Path(file_path).stem[:28] + "…" if len(Path(file_path).stem) > 28 else Path(file_path).stem
        sheet_name = "".join(c for c in sheet_name if c not in "[]:*?/\\") or "файл"
        # Уникализация на случай совпадения имён
        base = sheet_name
        suffix = 1
        while sheet_name in wb.sheetnames:
            suffix += 1
            sheet_name = f"{base}_{suffix}"[:31]

        ws = wb.create_sheet(sheet_name)
        ws.append(["Код", "Серьёзность", "Сообщение", "Расположение", "Что исправить"])
        for cell in ws[1]:
            cell.font = bold
            cell.fill = header_fill

        for v in violations:
            ws.append([v.check_code, v.severity, v.message, v.location, v.suggestion])

        for col_letter, width in zip("ABCDE", (10, 14, 60, 40, 60), strict=True):
            ws.column_dimensions[col_letter].width = width
        for row in ws.iter_rows(min_row=2):
            for cell in row:
                cell.alignment = wrap

    wb.save(str(output))


def _write_report(
    results: dict[str, list[Violation]], output: Path, profile_id: str
) -> str:
    """Выбрать формат отчёта по расширению файла. Возвращает подпись формата."""
    suffix = output.suffix.lower()
    if suffix == ".xlsx":
        _write_xlsx_report(results, output, profile_id)
        return "Excel"
    if suffix in {".md", ".markdown", ""}:
        _write_markdown_report(results, output, profile_id)
        return "Markdown"
    # Неизвестное расширение — fallback на Markdown
    _write_markdown_report(results, output, profile_id)
    return "Markdown"


@click.group()
@click.version_option(__version__, prog_name="gostforge")
def main() -> None:
    """gostforge — конструктор и нормоконтролёр документов по ГОСТу."""


@main.command()
@click.argument("path", type=click.Path(exists=True, path_type=Path))
@click.option("--profile", "-p", default="gost-7.32-2017", help="ID профиля для проверки")
@click.option(
    "--report",
    "-r",
    type=click.Path(path_type=Path),
    help="Путь к отчёту. Формат определяется по расширению: .xlsx → Excel, .md → Markdown",
)
@click.option("--quiet", "-q", is_flag=True, help="Показать только сводку и список кодов")
@click.option(
    "--no-record",
    is_flag=True,
    help="Не записывать результаты проверки в локальную БД истории.",
)
def check(
    path: Path, profile: str, report: Path | None, quiet: bool, no_record: bool
) -> None:
    """Проверить документ или папку документов на соответствие профилю."""
    try:
        prof = load_profile(profile)
    except FileNotFoundError as e:
        click.echo(f"Ошибка: {e}", err=True)
        sys.exit(2)

    targets = [path] if path.is_file() else sorted(path.glob("*.docx"))
    if not targets:
        click.echo(f"Не найдено .docx файлов в {path}", err=True)
        sys.exit(1)

    # Сводка по проверкам: сколько включено в профиле, сколько реально есть в реестре
    enabled_codes = {c for c, cfg in prof.checks.items() if cfg.enabled}
    available = set(registered_checks())
    runnable = enabled_codes & available
    skipped = enabled_codes - available

    click.echo(
        click.style("Профиль: ", bold=True) + profile + f" (v{prof.version})"
    )
    click.echo(
        f"Будет запущено проверок: {len(runnable)} из {len(enabled_codes)} включённых"
    )
    if skipped:
        click.echo(
            click.style("  Не реализованы: ", fg="yellow")
            + ", ".join(sorted(skipped))
        )

    results: dict[str, list[Violation]] = {}
    total_errors = 0
    start = time.perf_counter()

    for target in targets:
        document = parse_docx(target)
        violations = validate(document, prof)
        results[str(target)] = violations
        total_errors += sum(1 for v in violations if v.severity == "error")
        _print_violations(target, violations, quiet)

    elapsed = time.perf_counter() - start
    click.echo(
        "\n"
        + click.style("Итого: ", bold=True)
        + f"{len(targets)} файл(ов), {sum(len(v) for v in results.values())} нарушений "
        + f"за {elapsed:.2f} с"
    )

    if not no_record:
        _record_check_results(results, profile)

    if report:
        fmt = _write_report(results, report, profile)
        click.echo(click.style(f"{fmt}-отчёт сохранён: {report}", fg="green"))

    if total_errors > 0:
        sys.exit(1)


def _record_check_results(
    results: dict[str, list[Violation]], profile_id: str
) -> None:
    """Сохранить результаты прогона в локальную БД истории.

    Ошибки БД (нет места, нет прав) логируются и проглатываются —
    они не должны ломать основной workflow проверки.
    """
    try:
        from gostforge.db import get_connection, record_submission
    except ImportError:
        return
    try:
        with get_connection() as conn:
            for target, violations in results.items():
                record_submission(
                    conn,
                    filename=Path(target).name,
                    profile_id=profile_id,
                    violations=violations,
                )
    except Exception as exc:  # pragma: no cover — не валим CLI на БД
        click.echo(
            click.style(f"Не удалось записать в БД истории: {exc}", fg="yellow"),
            err=True,
        )


def _print_fixes(applied: list[FixApplied]) -> None:
    """Вывести таблицу применённых автоправок."""
    if not applied:
        click.echo("  " + click.style("[OK]", fg="green", bold=True) + "  Ничего исправлять не пришлось")
        return

    click.echo(
        "  "
        + click.style("[FIX]", fg="cyan", bold=True)
        + f"  Применено правок: {len(applied)}"
    )
    for record in applied:
        code = click.style(record.fixer_code, bold=True)
        click.echo(f"    {code}  {record.description}")
        if record.location:
            click.echo("       " + click.style(record.location, fg="bright_black"))


@main.command("fix")
@click.argument("path", type=click.Path(exists=True, path_type=Path))
@click.option(
    "-o",
    "--output",
    type=click.Path(path_type=Path),
    required=True,
    help="Куда сохранить исправленный .docx.",
)
@click.option("--profile", "-p", default="gost-7.32-2017")
@click.option(
    "--only",
    multiple=True,
    help="Применить только указанные коды (можно указать несколько раз).",
)
@click.option(
    "--dry-run",
    is_flag=True,
    help="Не записывать файл, только показать какие правки были бы применены.",  # noqa: RUF001
)
def fix_cmd(
    path: Path, output: Path, profile: str, only: tuple[str, ...], dry_run: bool
) -> None:
    """Применить безопасные автоисправления к .docx и записать результат в OUTPUT."""
    try:
        prof = load_profile(profile)
    except FileNotFoundError as e:
        click.echo(f"Ошибка: {e}", err=True)
        sys.exit(2)

    document = parse_docx(path)
    codes = list(only) if only else None
    applied = run_fix(document, prof, codes=codes)

    click.secho(f"\n>>> {path.name}", bold=True)
    _print_fixes(applied)

    if dry_run:
        click.echo(click.style("\n--dry-run: файл не записан.", fg="yellow"))
        return

    # Передаём source_docx=path, чтобы рисунки из исходного .docx
    # переносились в выходной как реальные изображения (а не placeholder).
    export_docx(document, prof, output, source_docx=path)
    click.echo(click.style(f"\nИсправленный документ сохранён: {output}", fg="green"))  # noqa: RUF001


@main.group()
def profiles() -> None:
    """Управление профилями."""


@profiles.command("list")
def profiles_list() -> None:
    """Показать доступные профили (builtin + установленные локально)."""
    from gostforge.profile import is_custom_profile

    for profile_id in list_profiles():
        if is_custom_profile(profile_id):
            click.echo(f"{profile_id}  " + click.style("[custom]", fg="green"))
        else:
            click.echo(f"{profile_id}  " + click.style("[builtin]", fg="cyan"))


@profiles.command("show")
@click.argument("profile_id")
def profiles_show(profile_id: str) -> None:
    """Показать содержимое профиля."""
    try:
        prof = load_profile(profile_id)
    except FileNotFoundError as e:
        click.echo(f"Ошибка: {e}", err=True)
        sys.exit(2)
    click.echo(prof.model_dump_json(indent=2))


@profiles.command("install")
@click.argument("path", type=click.Path(exists=True, path_type=Path))
@click.option(
    "--overwrite",
    is_flag=True,
    help="Перезаписать, если профиль с таким id уже установлен.",
)
def profiles_install(path: Path, overwrite: bool) -> None:
    """Установить YAML-профиль в локальный реестр.

    После установки профиль доступен всем командам по своему id
    (gostforge check ... --profile <id>, REST API и т.д.). YAML
    валидируется до записи в БД — неверная схема отвергается с
    понятной ошибкой.
    """
    try:
        from gostforge.db import get_connection, install_profile
    except ImportError as exc:  # pragma: no cover — db в stdlib
        click.echo(f"Ошибка импорта модуля БД: {exc}", err=True)
        sys.exit(2)

    yaml_content = path.read_text(encoding="utf-8")
    try:
        with get_connection() as conn:
            rec = install_profile(
                conn,
                yaml_content=yaml_content,
                source=str(path.resolve()),
                overwrite=overwrite,
            )
    except ValueError as exc:
        click.echo(click.style(f"Ошибка: {exc}", fg="red"), err=True)
        sys.exit(2)

    click.echo(
        click.style("Профиль установлен:", fg="green", bold=True)
        + f" {rec.profile_id}  ({rec.name}, v{rec.version})"
    )
    click.echo(f"  Источник:    {rec.source}")
    click.echo(f"  Установлен:  {rec.installed_at}")
    click.echo(
        "\nИспользовать: gostforge check FILE.docx --profile "
        + rec.profile_id
    )


@profiles.command("uninstall")
@click.argument("profile_id")
def profiles_uninstall(profile_id: str) -> None:
    """Удалить custom-профиль из локального реестра.

    Builtin-профили (gost-7.32-2017 и т. п.) удалить нельзя — они
    лежат в каталоге пакета, не в БД. Команда даст ошибку, если
    профиль не установлен локально.
    """
    try:
        from gostforge.db import get_connection, uninstall_profile
    except ImportError as exc:  # pragma: no cover
        click.echo(f"Ошибка импорта модуля БД: {exc}", err=True)
        sys.exit(2)

    with get_connection() as conn:
        removed = uninstall_profile(conn, profile_id)
    if removed:
        click.echo(
            click.style(f"Профиль удалён: {profile_id}", fg="green", bold=True)
        )
    else:
        click.echo(
            click.style(
                f"Профиль {profile_id!r} не установлен в локальный реестр.",
                fg="yellow",
            ),
            err=True,
        )
        sys.exit(1)


@profiles.command("validate")
@click.argument("path", type=click.Path(exists=True, path_type=Path))
def profiles_validate(path: Path) -> None:
    """Проверить YAML-файл профиля на корректность.

    Выводит сводку: число включённых проверок, сколько из них
    реализовано в текущем gostforge, нет ли ссылок на отсутствующие коды.
    """
    import yaml
    from gostforge.profile.schema import Profile

    try:
        data = yaml.safe_load(path.read_text(encoding="utf-8"))
    except yaml.YAMLError as e:
        click.echo(f"Ошибка парсинга YAML: {e}", err=True)
        sys.exit(2)

    try:
        prof = Profile(**data)
    except Exception as e:  # noqa: BLE001 — отображаем pydantic-ошибку как есть
        click.echo(f"Профиль не прошёл валидацию схемы:", err=True)
        click.echo(str(e), err=True)
        sys.exit(2)

    enabled_codes = {c for c, cfg in prof.checks.items() if cfg.enabled}
    available = set(registered_checks())
    runnable = enabled_codes & available
    skipped = enabled_codes - available

    click.echo(click.style(f"Профиль: {prof.id} v{prof.version}", bold=True))
    click.echo(f"  Включено проверок: {len(enabled_codes)}")
    click.echo(f"  Будет запущено: {len(runnable)}")
    if skipped:
        click.echo(
            click.style(f"  Не реализованы ({len(skipped)}): ", fg="yellow")
            + ", ".join(sorted(skipped))
        )
    if prof.extends:
        click.echo(f"  Наследует от: {prof.extends}")
    click.echo(click.style("[OK]", fg="green", bold=True) + " Файл валиден")


@profiles.command("diff")
@click.argument("profile_a")
@click.argument("profile_b")
def profiles_diff(profile_a: str, profile_b: str) -> None:
    """Сравнить два профиля по списку проверок и стилям.

    Выводит:
    - какие проверки включены только в A, только в B;
    - различия в параметрах общих проверок;
    - различия в основных полях styles.page/body.
    """
    try:
        a = load_profile(profile_a)
        b = load_profile(profile_b)
    except FileNotFoundError as e:
        click.echo(f"Ошибка: {e}", err=True)
        sys.exit(2)

    a_enabled = {c for c, cfg in a.checks.items() if cfg.enabled}
    b_enabled = {c for c, cfg in b.checks.items() if cfg.enabled}

    only_a = sorted(a_enabled - b_enabled)
    only_b = sorted(b_enabled - a_enabled)
    common = sorted(a_enabled & b_enabled)

    click.echo(click.style(f"=== {profile_a} vs {profile_b} ===", bold=True))

    if only_a:
        click.echo(click.style(f"\nТолько в {profile_a} ({len(only_a)}):", fg="cyan"))
        click.echo("  " + ", ".join(only_a))
    if only_b:
        click.echo(click.style(f"\nТолько в {profile_b} ({len(only_b)}):", fg="cyan"))
        click.echo("  " + ", ".join(only_b))

    # Отличия в параметрах общих проверок
    param_diffs = []
    for code in common:
        pa = a.checks[code].params
        pb = b.checks[code].params
        if pa != pb:
            param_diffs.append((code, pa, pb))
    if param_diffs:
        click.echo(click.style(f"\nРазные параметры ({len(param_diffs)}):", fg="yellow"))
        for code, pa, pb in param_diffs:
            click.echo(f"  {code}:")
            click.echo(f"    {profile_a}: {pa}")
            click.echo(f"    {profile_b}: {pb}")

    # Стили
    style_diffs = []
    if a.styles.page.margins_mm != b.styles.page.margins_mm:
        style_diffs.append(("margins_mm", a.styles.page.margins_mm, b.styles.page.margins_mm))
    for attr in ("font", "size_pt", "line_spacing", "first_line_indent_cm", "alignment"):
        va = getattr(a.styles.body, attr)
        vb = getattr(b.styles.body, attr)
        if va != vb:
            style_diffs.append((f"body.{attr}", va, vb))
    if style_diffs:
        click.echo(click.style(f"\nРазные стили ({len(style_diffs)}):", fg="yellow"))
        for name, va, vb in style_diffs:
            click.echo(f"  {name}: {profile_a}={va}, {profile_b}={vb}")

    if not only_a and not only_b and not param_diffs and not style_diffs:
        click.echo(click.style("\n[OK] Профили идентичны", fg="green"))


@main.command("checks")
def checks_list() -> None:
    """Показать список зарегистрированных проверок."""
    for code in registered_checks():
        click.echo(code)


@main.group()
def plugins() -> None:
    """Управление пользовательскими плагинами проверок."""


@plugins.command("list")
def plugins_list() -> None:
    """Показать загруженные плагины и предоставленные ими проверки."""
    from gostforge.plugins import (
        discover_plugin_files,
        load_plugins,
        plugins_dir,
    )
    from gostforge.validator.engine import _registry

    directory = plugins_dir()
    click.echo(f"Директория плагинов: {directory}")
    if not directory.exists():
        click.echo("  (не существует — создайте директорию для добавления плагинов)")
        return

    # Снимем снапшот текущего реестра, чтобы понять, какие проверки
    # добавились именно сейчас при загрузке плагинов.
    before = set(_registry)
    files = discover_plugin_files()
    load_plugins()
    after = set(_registry)
    added_codes = sorted(after - before)

    if not files:
        click.echo("  (плагинов не найдено)")
        return

    click.echo(f"\nНайдено файлов: {len(files)}")
    for f in files:
        click.echo(f"  - {f.name}")

    if added_codes:
        click.echo(f"\nДобавлены проверки: {', '.join(added_codes)}")


@plugins.command("dir")
def plugins_dir_cmd() -> None:
    """Показать путь к директории плагинов (или создать её)."""
    from gostforge.plugins import plugins_dir

    d = plugins_dir()
    if not d.exists():
        d.mkdir(parents=True, exist_ok=True)
        click.echo(f"Создана: {d}")
    else:
        click.echo(f"{d}")


@main.command()
@click.option("--port", default=8501, help="Порт (по умолчанию 8501).")
@click.option("--host", default="localhost", help="Адрес для bind (по умолчанию localhost).")
def ui(host: str, port: int) -> None:
    """Запустить веб-интерфейс на Streamlit."""
    try:
        import streamlit  # noqa: F401
    except ImportError:
        click.echo(
            "Streamlit не установлен. Установите gostforge[ui]:\n"
            "  pip install -e \".[ui]\"",
            err=True,
        )
        sys.exit(2)

    # Импортируем модуль приложения, чтобы получить путь до его файла.
    # Импорт обёрнут в try/except: в нём, помимо streamlit, могут быть
    # ошибки конфигурации, которые имеет смысл показать.
    import gostforge.web.app as app_module

    app_path = app_module.__file__
    cmd = [
        "streamlit",
        "run",
        str(app_path),
        "--server.address",
        host,
        "--server.port",
        str(port),
    ]
    subprocess.run(cmd, check=False)


@main.command()
@click.option("--host", default="127.0.0.1", help="Адрес для bind (по умолчанию 127.0.0.1).")
@click.option("--port", default=8000, help="Порт (по умолчанию 8000).")
@click.option(
    "--reload", is_flag=True, help="Включить hot-reload для разработки."
)
def serve(host: str, port: int, reload: bool) -> None:
    """Запустить REST API gostforge на FastAPI/uvicorn.

    Опциональная зависимость: pip install -e ".[api]". Полный список
    endpoints — docs/phase-3-api-spec.md. По умолчанию слушает только
    127.0.0.1 (запуск из публичной сети — за reverse-proxy с auth).
    """
    try:
        import uvicorn  # noqa: F401
    except ImportError:
        click.echo(
            "FastAPI/uvicorn не установлены. Установите gostforge[api]:\n"
            '  pip install -e ".[api]"',
            err=True,
        )
        sys.exit(2)

    import uvicorn

    uvicorn.run(
        "gostforge.api.app:app",
        host=host,
        port=port,
        reload=reload,
    )


@main.command()
@click.option("--limit", "-n", default=20, help="Сколько последних записей показать.")
@click.option(
    "--filename", "-f", default=None, help="Фильтр по имени файла (точное совпадение)."
)
@click.option(
    "--id",
    "submission_id",
    type=int,
    default=None,
    help="Показать детали одного submission по id.",
)
def history(limit: int, filename: str | None, submission_id: int | None) -> None:
    """История проверок из локальной БД.

    Без параметров показывает последние ``--limit`` записей с
    summary по severity. С ``--id`` показывает детали конкретного
    submission со списком всех найденных нарушений.

    БД лежит в ``~/.gostforge/gostforge.db`` (переопределяется env
    ``GOSTFORGE_DB_PATH``) и автоматически создаётся при первом
    ``gostforge check``.
    """
    try:
        from gostforge.db import get_connection, get_submission, list_submissions
    except ImportError as exc:  # pragma: no cover — db в stdlib
        click.echo(f"Не удалось импортировать модуль БД: {exc}", err=True)
        sys.exit(2)

    with get_connection() as conn:
        if submission_id is not None:
            sub = get_submission(conn, submission_id)
            if sub is None:
                click.echo(f"Submission #{submission_id} не найден.", err=True)
                sys.exit(1)
            _print_submission_details(sub)
            return

        items = list_submissions(conn, limit=limit, filename=filename)
        if not items:
            click.echo("История пуста. Запустите 'gostforge check ...' для первой записи.")
            return

        click.echo(click.style(f"Последние {len(items)} проверок:\n", bold=True))
        for s in items:
            severity_str = (
                f"{click.style(str(s.error_count), fg='red')}e/"
                f"{click.style(str(s.warning_count), fg='yellow')}w/"
                f"{click.style(str(s.info_count), fg='cyan')}i"
            )
            click.echo(
                f"  #{s.id:>4}  {s.created_at}  {severity_str}  "
                f"[{s.profile_id}]  {s.filename}"
            )
        click.echo(
            "\nДля деталей: gostforge history --id <N>"
        )


@main.group()
def comment() -> None:
    """Комментарии к submission-ам (совместная работа)."""


def _default_author() -> str:
    """Имя автора по умолчанию — из env или getpass.getuser()."""
    env = os.environ.get("GOSTFORGE_DEFAULT_AUTHOR", "").strip()
    if env:
        return env
    try:
        import getpass

        return getpass.getuser()
    except Exception:  # pragma: no cover - graceful
        return ""


@comment.command("add")
@click.argument("submission_id", type=int)
@click.argument("body")
@click.option(
    "--author",
    default=None,
    help="Имя автора. Если не задано — берётся из env GOSTFORGE_DEFAULT_AUTHOR "
    "или getpass.getuser().",
)
@click.option(
    "--role",
    type=click.Choice(["student", "supervisor", "anonymous"]),
    default="anonymous",
    help="Роль автора (student/supervisor/anonymous).",
)
def comment_add(
    submission_id: int, body: str, author: str | None, role: str
) -> None:
    """Добавить комментарий к submission.

    Пример:

        gostforge comment add 42 "Переделай введение" --role supervisor

    Если submission_id не существует или body пустой — выходим с
    кодом 2 + сообщением.
    """
    from gostforge.db import add_comment, get_connection

    final_author = author if author is not None else _default_author()
    try:
        with get_connection() as conn:
            c = add_comment(
                conn,
                submission_id=submission_id,
                body=body,
                author=final_author,
                role=role,
            )
    except ValueError as exc:
        click.echo(click.style(f"Ошибка: {exc}", fg="red"), err=True)
        sys.exit(2)

    click.echo(
        click.style("Комментарий добавлен:", fg="green", bold=True)
        + f" #{c.id} к submission #{c.submission_id}"
    )
    click.echo(f"  Автор: {c.author or '—'}  [{c.role}]")
    click.echo(f"  Время: {c.created_at}")


@comment.command("list")
@click.argument("submission_id", type=int)
@click.option(
    "--unresolved",
    is_flag=True,
    help="Показать только незакрытые комментарии.",
)
def comment_list(submission_id: int, unresolved: bool) -> None:
    """Показать все комментарии к submission.

    Закрытые помечены ``✓``, открытые — ``●``.
    """
    from gostforge.db import get_connection, list_comments

    with get_connection() as conn:
        items = list_comments(
            conn,
            submission_id=submission_id,
            include_resolved=not unresolved,
        )
    if not items:
        click.echo("Комментариев нет.")
        return
    for c in items:
        role_color = {
            "supervisor": "magenta",
            "student": "blue",
            "anonymous": "bright_black",
        }.get(c.role, "white")
        role_label = click.style(f"[{c.role}]", fg=role_color)
        status = (
            click.style("✓", fg="green") if c.resolved else click.style("●", fg="yellow")
        )
        author = c.author or "—"
        click.echo(
            f"#{c.id} {status} {role_label} {author} "
            + click.style(c.created_at, fg="bright_black")
        )
        for line in c.body.splitlines():
            click.echo(f"    {line}")


@comment.command("resolve")
@click.argument("comment_id", type=int)
@click.option(
    "--reopen",
    is_flag=True,
    help="Снять отметку resolved (вернуть в открытые).",
)
def comment_resolve(comment_id: int, reopen: bool) -> None:
    """Пометить комментарий как resolved (или снять отметку через --reopen)."""
    from gostforge.db import get_connection, resolve_comment

    with get_connection() as conn:
        ok = resolve_comment(conn, comment_id, resolved=not reopen)
    if not ok:
        click.echo(
            click.style(f"Комментарий #{comment_id} не найден.", fg="yellow"),
            err=True,
        )
        sys.exit(1)
    action = "переоткрыт" if reopen else "закрыт"
    click.echo(
        click.style(f"Комментарий #{comment_id} {action}.", fg="green", bold=True)
    )


@comment.command("delete")
@click.argument("comment_id", type=int)
def comment_delete(comment_id: int) -> None:
    """Удалить комментарий."""
    from gostforge.db import delete_comment, get_connection

    with get_connection() as conn:
        ok = delete_comment(conn, comment_id)
    if not ok:
        click.echo(
            click.style(f"Комментарий #{comment_id} не найден.", fg="yellow"),
            err=True,
        )
        sys.exit(1)
    click.echo(
        click.style(f"Комментарий #{comment_id} удалён.", fg="green", bold=True)
    )


def _print_submission_details(sub: object) -> None:
    """Распечатать одну запись со списком violations и комментариев."""
    click.echo(click.style(f"Submission #{sub.id}", bold=True))  # type: ignore[attr-defined]
    click.echo(f"  Файл:      {sub.filename}")  # type: ignore[attr-defined]
    click.echo(f"  Профиль:   {sub.profile_id}")  # type: ignore[attr-defined]
    click.echo(f"  Время:     {sub.created_at}")  # type: ignore[attr-defined]
    click.echo(
        f"  Сводка:    "
        f"{click.style(str(sub.error_count), fg='red')} error / "  # type: ignore[attr-defined]
        f"{click.style(str(sub.warning_count), fg='yellow')} warning / "  # type: ignore[attr-defined]
        f"{click.style(str(sub.info_count), fg='cyan')} info"  # type: ignore[attr-defined]
    )
    if not sub.violations:  # type: ignore[attr-defined]
        click.echo("\n  Нарушений нет.")
    else:
        click.echo("\n  Нарушения:")
        for v in sub.violations:  # type: ignore[attr-defined]
            color = {"error": "red", "warning": "yellow", "info": "cyan"}.get(
                v.severity, "white"
            )
            click.echo(
                f"    {click.style(v.severity.upper(), fg=color)}  "
                f"{v.code:>6}  {v.message}"
            )
            if v.location:
                click.echo(
                    click.style(f"            {v.location}", fg="bright_black")
                )
            if v.suggestion:
                click.echo(click.style(f"          → {v.suggestion}", fg="green"))

    _print_submission_comments(sub.id)  # type: ignore[attr-defined]


def _print_submission_comments(submission_id: int) -> None:
    """Распечатать ленту комментариев к submission (если есть)."""
    try:
        from gostforge.db import get_connection, list_comments
    except ImportError:
        return
    try:
        with get_connection() as conn:
            comments = list_comments(conn, submission_id=submission_id)
    except Exception:
        return
    if not comments:
        return
    click.echo("\n  " + click.style("Комментарии:", bold=True))
    for c in comments:
        role_color = {
            "supervisor": "magenta",
            "student": "blue",
            "anonymous": "bright_black",
        }.get(c.role, "white")
        role_label = click.style(f"[{c.role}]", fg=role_color)
        status = (
            click.style(" ✓", fg="green") if c.resolved else click.style(" ●", fg="yellow")
        )
        author = c.author or "—"
        click.echo(
            f"    #{c.id}{status}  {role_label}  {author}  "
            + click.style(c.created_at, fg="bright_black")
        )
        for line in c.body.splitlines():
            click.echo(f"        {line}")


def _normalize_message(text: str) -> str:
    """Нормализовать текст сообщения для устойчивого сравнения нарушений.

    Сжимает пробелы, обрезает пробелы по краям и приводит к нижнему регистру.
    Это уменьшает ложные срабатывания «нарушение пропало / появилось» из-за
    малозначимых отличий в форматировании сообщений.
    """
    return " ".join(text.split()).strip().lower()


def _violation_fingerprint(v: Violation) -> tuple[str, str, str]:
    """Отпечаток нарушения для сравнения двух документов.

    Тройка `(check_code, location, normalize(message))` — компромисс между
    стабильностью (одна и та же ошибка в двух прогонах даёт один и тот же
    отпечаток) и уникальностью (две разные ошибки в одном месте отличаются
    нормализованным сообщением).
    """
    return (v.check_code, v.location, _normalize_message(v.message))


def _print_violations_brief(violations: list[Violation], indent: str = "  ") -> None:
    """Кратко вывести список violations с цветовой пометкой по серьёзности."""
    for v in violations:
        _, style, short = _SEVERITY_STYLE.get(v.severity, ("", {}, v.severity.upper()))
        tag = click.style(f"[{short}]", **style) if style else f"[{short}]"
        code = click.style(v.check_code, bold=True)
        click.echo(f"{indent}{tag} {code}  {v.message}")
        if v.location:
            click.echo(indent + "       " + click.style(v.location, fg="bright_black"))


@main.command()
@click.argument("file_a", type=click.Path(exists=True, path_type=Path))
@click.argument("file_b", type=click.Path(exists=True, path_type=Path))
@click.option("--profile", "-p", default="gost-7.32-2017", help="ID профиля для проверки")
@click.option("--quiet", "-q", is_flag=True, help="Не показывать детали по нарушениям")
def diff(file_a: Path, file_b: Path, profile: str, quiet: bool) -> None:
    """Сравнить два .docx по списку нарушений.

    Выводит:
    - какие нарушения появились в B (не было в A)
    - какие нарушения исчезли в A (нет в B)
    - сводку: было N нарушений → стало M

    Полезно для CI (стало ли нарушений меньше после правки), проверки
    эффекта `gostforge fix`, анализа версий «черновик → финал».

    Exit code: 0 если число error-нарушений в B не больше, чем в A;
    1 если в B появились новые error-нарушения (регрессия).
    """
    try:
        prof = load_profile(profile)
    except FileNotFoundError as e:
        click.echo(f"Ошибка: {e}", err=True)
        sys.exit(2)

    document_a = parse_docx(file_a)
    document_b = parse_docx(file_b)
    violations_a = validate(document_a, prof)
    violations_b = validate(document_b, prof)

    fingerprints_a = {_violation_fingerprint(v): v for v in violations_a}
    fingerprints_b = {_violation_fingerprint(v): v for v in violations_b}

    fixed_keys = set(fingerprints_a) - set(fingerprints_b)
    new_keys = set(fingerprints_b) - set(fingerprints_a)

    fixed = [fingerprints_a[k] for k in fixed_keys]
    introduced = [fingerprints_b[k] for k in new_keys]

    errors_a = sum(1 for v in violations_a if v.severity == "error")
    errors_b = sum(1 for v in violations_b if v.severity == "error")

    click.echo(
        click.style("Профиль: ", bold=True) + profile + f" (v{prof.version})"
    )
    click.echo(
        click.style("A: ", bold=True)
        + f"{file_a.name} — {len(violations_a)} нарушений ({errors_a} ошибок)"
    )
    click.echo(
        click.style("B: ", bold=True)
        + f"{file_b.name} — {len(violations_b)} нарушений ({errors_b} ошибок)"
    )

    if not fixed and not introduced:
        click.echo(click.style("\n[OK] Изменений в нарушениях нет.", fg="green", bold=True))
        sys.exit(0)

    if fixed:
        click.echo(
            "\n"
            + click.style(f"Исчезло нарушений: {len(fixed)}", fg="green", bold=True)
        )
        if not quiet:
            _print_violations_brief(fixed)

    if introduced:
        click.echo(
            "\n"
            + click.style(f"Появилось нарушений: {len(introduced)}", fg="red", bold=True)
        )
        if not quiet:
            _print_violations_brief(introduced)

    delta = errors_b - errors_a
    if delta > 0:
        click.echo(
            "\n"
            + click.style(
                f"Регрессия: число error-нарушений выросло ({errors_a} → {errors_b}, +{delta}).",
                fg="red",
                bold=True,
            )
        )
        sys.exit(1)
    elif delta < 0:
        click.echo(
            "\n"
            + click.style(
                f"Прогресс: число error-нарушений уменьшилось ({errors_a} → {errors_b}, {delta}).",
                fg="green",
                bold=True,
            )
        )
    else:
        click.echo(
            "\n"
            + click.style(
                f"Число error-нарушений не изменилось ({errors_a}).",
                fg="yellow",
            )
        )
    sys.exit(0)


@main.command()
@click.argument("path", type=click.Path(exists=True, path_type=Path))
def stats(path: Path) -> None:
    """Показать структурную статистику документа.

    Считает число разделов, параграфов, таблиц, рисунков, источников
    и слов. Не зависит от профиля и не выполняет проверки.
    """
    from gostforge.stats import compute_stats

    targets = [path] if path.is_file() else sorted(path.glob("*.docx"))
    if not targets:
        click.echo(f"Не найдено .docx файлов в {path}", err=True)
        sys.exit(1)

    for target in targets:
        document = parse_docx(target)
        s = compute_stats(document)
        click.secho(f"\n>>> {target.name}", bold=True)
        rows = [
            ("Секций вёрстки (PageSection)", s.page_sections),
            ("Разделов 1 уровня", s.logical_sections_level_1),
            ("Разделов всего", s.logical_sections_total),
            ("Параграфов всего", s.paragraphs),
            ("  …непустых", s.paragraphs_non_empty),
            ("Таблиц", s.tables),
            ("Рисунков", s.figures),
            ("Источников", s.bibliography_entries),
            ("Слов", s.words),
            ("Символов", s.characters),
        ]
        for label, value in rows:
            click.echo(f"  {label:<32} {value}")


@main.command()
@click.argument("path", type=click.Path(exists=True, path_type=Path))
@click.option(
    "-o",
    "--output",
    type=click.Path(path_type=Path),
    required=True,
    help="Куда сохранить .docx с inline-пометками.",
)
@click.option("--profile", "-p", default="gost-7.32-2017", help="ID профиля.")
@click.option(
    "--style",
    type=click.Choice(["inline", "comments"]),
    default="comments",
    show_default=True,
    help=(
        "Стиль аннотации: настоящие комментарии Word (comments) или "
        "inline-маркеры в тексте (inline)."
    ),
)
def annotate(path: Path, output: Path, profile: str, style: str) -> None:
    """Создать .docx с пометками о нарушениях.

    По умолчанию (``--style comments``) вставляются настоящие OOXML-комментарии
    Word: в документе создаётся часть ``word/comments.xml`` и в проблемных
    параграфах ставятся ``<w:commentRangeStart/End>`` + reference-run.
    Word и LibreOffice отображают это как боковые выноски.

    При ``--style inline`` используется старый режим: в начало проблемного
    параграфа вставляется маркер вида ``[F.01: <текст ошибки>]`` курсивом,
    красным цветом. Пометки уровня документа уходят в первый параграф.
    """
    from gostforge.annotator import annotate_docx
    try:
        prof = load_profile(profile)
    except FileNotFoundError as e:
        click.echo(f"Ошибка: {e}", err=True)
        sys.exit(2)
    n = annotate_docx(path, output, prof, style=style)  # type: ignore[arg-type]
    click.echo(f"Создано пометок: {n}")
    click.echo(f"Аннотированный документ сохранён: {output}")


@main.command()
@click.argument("path", type=click.Path(exists=True, path_type=Path))
@click.option(
    "-o",
    "--output",
    type=click.Path(path_type=Path),
    required=True,
    help="Куда сохранить .pdf.",
)
@click.option(
    "--timeout",
    type=float,
    default=60.0,
    help="Таймаут конвертации в секундах.",
)
def pdf(path: Path, output: Path, timeout: float) -> None:
    """Конвертировать .docx → .pdf через LibreOffice headless.

    Требует установленного LibreOffice. Полезно для генерации
    PDF-версии работы после применения автофиксов.
    """
    from gostforge.pdf_exporter import LibreOfficeNotFoundError, convert_to_pdf

    try:
        result = convert_to_pdf(path, output, timeout=timeout)
    except LibreOfficeNotFoundError as e:
        click.echo(f"Ошибка: {e}", err=True)
        sys.exit(3)
    except subprocess.TimeoutExpired:
        click.echo(
            f"Конвертация прервана по таймауту ({timeout}s)", err=True
        )
        sys.exit(4)
    except subprocess.CalledProcessError as e:
        click.echo(
            f"LibreOffice вернул ошибку (код {e.returncode}):", err=True
        )
        stderr = e.stderr or b""
        click.echo(stderr.decode("utf-8", errors="replace"), err=True)
        sys.exit(5)
    click.echo(f"PDF сохранён: {result}")


@main.command()
@click.argument("output", type=click.Path(path_type=Path))
@click.option(
    "--template",
    "-t",
    type=click.Choice(["coursework", "bachelor_thesis", "research_report"]),
    default="coursework",
    help="Шаблон работы.",
)
@click.option("--title", required=True, help="Название работы.")
@click.option("--author", default="", help="Автор.")
@click.option("--supervisor", default="", help="Руководитель.")
@click.option("--organization", default="", help="Организация.")
@click.option("--year", type=int, default=None, help="Год.")
@click.option("--profile", "-p", default="gost-7.32-2017", help="Профиль ГОСТа.")
def new(
    output: Path,
    template: str,
    title: str,
    author: str,
    supervisor: str,
    organization: str,
    year: int | None,
    profile: str,
) -> None:
    """Создать новую работу из шаблона.

    Пример: gostforge new my-coursework.docx --template coursework
        --title "Курсовая по нормоконтролю" --author "Иванов И. И." --year 2026
    """
    from gostforge.builder.templates import (
        bachelor_thesis_template,
        coursework_template,
        research_report_template,
    )

    if template == "coursework":
        builder = coursework_template(
            title=title,
            author=author,
            supervisor=supervisor,
            organization=organization,
            year=year,
        )
    elif template == "bachelor_thesis":
        builder = bachelor_thesis_template(
            title=title,
            author=author,
            supervisor=supervisor,
            organization=organization,
            year=year,
        )
    else:  # research_report
        builder = research_report_template(
            title=title,
            year=year,
            organization=organization,
        )

    try:
        builder.save(output, profile=profile)
    except FileNotFoundError as e:
        click.echo(f"Профиль не найден: {e}", err=True)
        sys.exit(2)
    except ValueError as e:
        click.echo(f"Документ не прошёл валидацию: {e}", err=True)
        sys.exit(3)

    click.echo(f"Создан файл: {output}")
    click.echo("Откройте его в Word / LibreOffice и заполните плейсхолдеры.")


@main.command("new-state")
@click.option(
    "--template",
    type=click.Choice(["coursework", "thesis", "research_report", "empty"]),
    default="empty",
    help="Тип шаблона (по умолчанию пустой каркас).",
)
@click.option(
    "--title", type=str, default="Новая работа", help="Название работы."
)
@click.option("--author", type=str, default="", help="Автор.")
@click.option(
    "--year",
    type=int,
    default=None,
    help="Год работы (по умолчанию текущий).",
)
@click.option(
    "--profile", "profile_id", type=str, default="gost-7.32-2017",
    help="Идентификатор профиля.",
)
@click.option(
    "-o",
    "--output",
    type=click.Path(path_type=Path),
    required=True,
    help="Куда сохранить JSON-state.",
)
def new_state_cmd(
    template: str,
    title: str,
    author: str,
    year: int | None,
    profile_id: str,
    output: Path,
) -> None:
    """Создать пустой JSON-state для конструктора по выбранному шаблону.

    Зеркало команды ``gostforge new`` (которая создаёт .docx), но
    результат — JSON для UI/CLI-цикла. Использует существующие
    шаблоны из ``gostforge.builder.templates``.

    Полный цикл: создать → отредактировать → собрать::

        gostforge new-state --template coursework --title "Анализ X" -o state.json
        gostforge ui    # правим в Streamlit
        gostforge generate state.json -o work.docx
    """
    from datetime import date  # noqa: PLC0415

    from gostforge.builder.templates import (  # noqa: PLC0415
        bachelor_thesis_template,
        coursework_template,
        research_report_template,
    )
    from gostforge.web.builder_editor import document_to_state  # noqa: PLC0415

    if year is None:
        year = date.today().year

    if template == "empty":
        # Минимальный каркас: один раздел «Введение» + список источников.
        state = {
            "title": title,
            "author": author,
            "year": year,
            "profile_id": profile_id,
            "sections": [
                {
                    "heading": "Введение",
                    "blocks": [
                        {"kind": "paragraph", "text": ""}
                    ],
                },
                {
                    "heading": "Список использованных источников",
                    "is_bibliography": True,
                    "references": [],
                },
            ],
        }
    else:
        # Используем существующий шаблон builder-а, собираем Document
        # и конвертируем обратно в state. Это даёт ровно то же
        # содержимое, что и `gostforge new --template=X`, но в формате
        # для конструктора.
        if template == "coursework":
            builder = coursework_template(
                title=title, author=author, year=year
            )
        elif template == "thesis":
            builder = bachelor_thesis_template(
                title=title, author=author, year=year
            )
        else:  # research_report
            builder = research_report_template(title=title, year=year)
        document = builder.build()
        document.profile_id = profile_id
        state = document_to_state(document)

    output.write_text(
        json.dumps(state, ensure_ascii=False, indent=2),
        encoding="utf-8",
    )
    n = len(state.get("sections", []))
    click.echo(
        f"Создан {output} ({n} разделов, шаблон '{template}'). "
        f"Откройте в `gostforge ui` или редактируйте JSON напрямую."
    )


@main.command("generate")
@click.argument("state_path", type=click.Path(exists=True, path_type=Path))
@click.option(
    "-o",
    "--output",
    type=click.Path(path_type=Path),
    required=True,
    help="Куда сохранить .docx.",
)
@click.option(
    "--profile",
    "profile_override",
    type=str,
    default=None,
    help="Профиль для экспорта. По умолчанию — из state.json (profile_id).",
)
def generate_cmd(
    state_path: Path, output: Path, profile_override: str | None
) -> None:
    """Сгенерировать .docx из JSON-state конструктора.

    Это зеркало команды ``import-docx``: вместе они образуют полный
    CLI-цикл:

    \b
        gostforge import-docx work.docx -o state.json
        # ... редактируете state.json вручную или скриптом ...
        gostforge generate state.json -o new.docx

    Формат state.json — тот же, что в Streamlit-конструкторе
    (sidebar → «Загрузить сохранение»). Профиль берётся из
    ``state.profile_id``; ``--profile`` его переопределяет.
    """
    from gostforge.web.builder_editor import (  # noqa: PLC0415
        _build_document_from_state,
    )

    try:
        state = json.loads(state_path.read_text(encoding="utf-8"))
    except json.JSONDecodeError as exc:
        click.echo(f"Не удалось прочитать JSON: {exc}", err=True)
        sys.exit(2)
    if not isinstance(state, dict) or "sections" not in state:
        click.echo("В state.json должен быть ключ 'sections'.", err=True)
        sys.exit(2)
    if profile_override:
        state = dict(state)
        state["profile_id"] = profile_override

    try:
        data = _build_document_from_state(state)
    except FileNotFoundError as exc:
        click.echo(f"Профиль не найден: {exc}", err=True)
        sys.exit(2)
    except ValueError as exc:
        click.echo(f"Ошибка сборки: {exc}", err=True)
        sys.exit(3)

    output.write_bytes(data)
    click.echo(f"Сгенерирован файл: {output}")


@main.command("export-md")
@click.argument("state_path", type=click.Path(exists=True, path_type=Path))
@click.option(
    "-o",
    "--output",
    type=click.Path(path_type=Path),
    required=True,
    help="Куда сохранить .md.",
)
def export_md_cmd(state_path: Path, output: Path) -> None:
    """Экспортировать state-конструктора в Markdown.

    Удобно для:
    * быстрого ревью работы в git/pull-request;
    * конвертации в другие форматы через pandoc;
    * интеграции с системами на базе Markdown (Obsidian, Notion).

    Маппинг:
    * sections с level=1 → ``# Заголовок``;
    * sub (level=2) → ``## Заголовок``;
    * subsub (level=3) → ``### Заголовок``;
    * paragraph → обычный абзац;
    * list/ordered → ``1. item`` / ``- item``;
    * table → GFM-таблица с caption-строкой;
    * figure → ``![caption](image_path)``;
    * formula → ``$$ latex $$``;
    * references → нумерованный список.

    Не сохраняет: disabled_checks, profile-id (это metadata
    конструктора, не имеет смысла в Markdown).
    """
    try:
        state = json.loads(state_path.read_text(encoding="utf-8"))
    except json.JSONDecodeError as exc:
        click.echo(f"Не удалось прочитать JSON: {exc}", err=True)
        sys.exit(2)
    if not isinstance(state, dict) or "sections" not in state:
        click.echo("В state.json должен быть ключ 'sections'.", err=True)
        sys.exit(2)
    md = _state_to_markdown(state)
    output.write_text(md, encoding="utf-8")
    click.echo(f"Создан {output} ({len(md.splitlines())} строк)")


def _state_to_markdown(state: dict[str, Any]) -> str:
    """Сериализовать state в строку Markdown (GFM).

    Используется CLI export-md, а также может быть переиспользована
    в UI как preview перед скачиванием.
    """
    lines: list[str] = []
    title = state.get("title", "").strip()
    if title:
        lines.append(f"# {title}")
        lines.append("")
        meta_parts = []
        if state.get("author"):
            meta_parts.append(f"**Автор:** {state['author']}")
        if state.get("year"):
            meta_parts.append(f"**Год:** {state['year']}")
        if state.get("supervisor"):
            meta_parts.append(f"**Руководитель:** {state['supervisor']}")
        if meta_parts:
            lines.append(" · ".join(meta_parts))
            lines.append("")

    for sec in state.get("sections", []):
        _section_to_md(sec, depth=2, lines=lines)
        lines.append("")
    return "\n".join(lines).rstrip() + "\n"


def _section_to_md(
    section: dict[str, Any], *, depth: int, lines: list[str]
) -> None:
    """Записать один раздел и его потомков в lines.

    ``depth`` — уровень заголовка (1=#, 2=##, ...). Top-level разделы
    идут с depth=2, потому что depth=1 зарезервирован под title работы.
    """
    heading = section.get("heading", "").strip() or "(без названия)"
    hash_prefix = "#" * min(depth, 6)
    lines.append(f"{hash_prefix} {heading}")
    lines.append("")

    if section.get("is_bibliography"):
        for i, ref in enumerate(section.get("references", []) or [], start=1):
            lines.append(f"{i}. {ref}")
        lines.append("")
        return

    for block in section.get("blocks") or []:
        _block_to_md(block, lines=lines)

    for sub in section.get("subsections") or []:
        _section_to_md(sub, depth=depth + 1, lines=lines)


def _paragraph_to_md(block: dict[str, Any]) -> str:
    """Сериализовать параграф (text или runs) в Markdown-строку.

    Если block имеет 'runs' с bold/italic — оборачиваем фрагменты:
    * bold + italic → '***x***'
    * bold → '**x**'
    * italic → '*x*'
    * underline (MD не поддерживает нативно) → '_x_' (italic-fallback,
      без потерь читаемости).

    Для простого text= возвращаем как есть.
    """
    text = block.get("text", "")
    runs = block.get("runs") or []
    if not runs:
        return text
    out_parts: list[str] = []
    for r in runs:
        if r.get("kind") != "text":
            # Для xref/citation/formula пока берём текст «как есть»
            # (формулы → $...$, xref → [текст], citation → [N]).
            if r.get("kind") == "formula":
                out_parts.append(f"${r.get('latex', '')}$")
            elif r.get("kind") == "citation":
                sid = r.get("source_id", "")
                page = r.get("page", "")
                out_parts.append(
                    f"[{sid}, с. {page}]" if page else f"[{sid}]"
                )
            elif r.get("kind") == "xref":
                # xref в Markdown ёще нет нативного аналога — кладём
                # placeholder. import-md его обратно не превратит,
                # но текст не потеряется.
                tgt = r.get("target_id", "")
                prefix = r.get("prefix", "")
                out_parts.append(f"{prefix}[→{tgt}]")
            continue
        t = r.get("text", "")
        if not t:
            continue
        bold = bool(r.get("bold"))
        italic = bool(r.get("italic"))
        if bold and italic:
            out_parts.append(f"***{t}***")
        elif bold:
            out_parts.append(f"**{t}**")
        elif italic:
            out_parts.append(f"*{t}*")
        else:
            out_parts.append(t)
    return "".join(out_parts)


def _block_to_md(block: dict[str, Any], *, lines: list[str]) -> None:
    """Сериализовать один Block в Markdown-строки.

    Параграфы с rich-runs (bold/italic) экспортируются с
    Markdown-разметкой: bold=True → ``**текст**``, italic=True →
    ``*текст*``, bold+italic → ``***текст***``. Это даёт настоящий
    rich-Markdown при export-md и round-trip через import-md.
    """
    kind = block.get("kind", "")
    if kind == "paragraph":
        text = _paragraph_to_md(block)
        if text.strip():
            lines.append(text)
            lines.append("")
    elif kind == "list":
        items = block.get("items") or []
        ordered = block.get("ordered", False)
        for i, item in enumerate(items, start=1):
            marker = f"{i}." if ordered else "-"
            lines.append(f"{marker} {item}")
        lines.append("")
    elif kind == "table":
        headers = block.get("headers") or []
        rows = block.get("rows") or []
        caption = block.get("caption", "")
        if caption:
            lines.append(f"**{caption}**")
            lines.append("")
        if headers:
            lines.append("| " + " | ".join(str(h) for h in headers) + " |")
            lines.append("|" + "|".join("---" for _ in headers) + "|")
        for row in rows:
            lines.append("| " + " | ".join(str(c) for c in row) + " |")
        lines.append("")
    elif kind == "figure":
        path = block.get("image_path", "")
        caption = block.get("caption", "")
        if path:
            lines.append(f"![{caption}]({path})")
        elif caption:
            lines.append(f"*Рисунок: {caption}*")
        lines.append("")
    elif kind == "formula":
        latex = block.get("latex", "")
        if latex:
            lines.append(f"$$ {latex} $$")
            lines.append("")


@main.command("stats-state")
@click.argument("state_path", type=click.Path(exists=True, path_type=Path))
@click.option(
    "--json", "as_json", is_flag=True,
    help="Вывод в JSON (для скриптов). По умолчанию — таблица.",
)
def stats_state_cmd(state_path: Path, as_json: bool) -> None:
    """Числовые метрики state-файла конструктора.

    Считает: разделов, параграфов, таблиц, рисунков, формул, элементов
    списков, источников, общее число слов, оценочное число знаков.
    Полезно для быстрой проверки прогресса работы без открытия UI.

    Пример::

        gostforge stats-state draft.json
        gostforge stats-state draft.json --json | jq .total_words
    """
    try:
        state = json.loads(state_path.read_text(encoding="utf-8"))
    except json.JSONDecodeError as exc:
        click.echo(f"Не удалось прочитать JSON: {exc}", err=True)
        sys.exit(2)
    if not isinstance(state, dict) or "sections" not in state:
        click.echo("В state.json должен быть ключ 'sections'.", err=True)
        sys.exit(2)

    from gostforge.web.builder_editor import (  # noqa: PLC0415
        _compute_progress_metrics,
    )

    metrics = _compute_progress_metrics(state)
    if as_json:
        click.echo(json.dumps(metrics, ensure_ascii=False, indent=2))
        return

    click.echo(f"Состояние: {state_path}")
    click.echo(f"  Заголовок: {state.get('title', '(без названия)')}")
    click.echo(f"  Автор: {state.get('author') or '(не указан)'}")
    click.echo(f"  Год: {state.get('year', '(не указан)')}")
    click.echo(f"  Профиль: {state.get('profile_id', 'gost-7.32-2017')}")
    click.echo("")
    click.echo("Прогресс:")
    click.echo(f"  Разделов: {metrics['sections_filled']}/{metrics['sections_total']}")
    click.echo(f"  Параграфов: {metrics['paragraphs_nonempty']}/{metrics['paragraphs_total']}")
    click.echo(f"  Таблиц: {metrics['tables']}")
    click.echo(f"  Рисунков: {metrics['figures']}")
    click.echo(f"  Формул: {metrics['formulas']}")
    click.echo(f"  Элементов списков: {metrics['list_items']}")
    click.echo(f"  Источников: {metrics['references_total']}")
    click.echo("")
    click.echo("Объём:")
    click.echo(f"  Слов: {metrics['total_words']}")
    # Оценка знаков: ~6 символов на слово (среднее для русского).
    est_chars = metrics["total_words"] * 6
    click.echo(f"  Знаков (≈): {est_chars}")


@main.command("check-state")
@click.argument("state_path", type=click.Path(exists=True, path_type=Path))
@click.option(
    "--json", "as_json", is_flag=True,
    help="Вывод в JSON (для скриптов).",
)
@click.option(
    "--profile",
    "profile_override",
    type=str,
    default=None,
    help="Профиль для проверки. По умолчанию — из state.profile_id.",
)
def check_state_cmd(
    state_path: Path, as_json: bool, profile_override: str | None
) -> None:
    """Прогон нормоконтроля над state-файлом без сохранения в .docx.

    Использует тот же путь, что и live-нормоконтроль в UI: state →
    builder → Document → validate. В разы быстрее чем
    `generate + check`, потому что не пишет промежуточный .docx.

    Учитывает ``disabled_checks`` каждого раздела — нарушения для
    отключённых проверок отфильтровываются.

    Exit code: 0 — нарушений нет; 1 — найдены ошибки; 2 — невалидный
    state JSON или профиль не найден.
    """
    try:
        state = json.loads(state_path.read_text(encoding="utf-8"))
    except json.JSONDecodeError as exc:
        click.echo(f"Не удалось прочитать JSON: {exc}", err=True)
        sys.exit(2)
    if not isinstance(state, dict) or "sections" not in state:
        click.echo("В state.json должен быть ключ 'sections'.", err=True)
        sys.exit(2)
    if profile_override:
        state = dict(state)
        state["profile_id"] = profile_override

    from gostforge.web.builder_editor import (  # noqa: PLC0415
        _compute_live_validation_summary,
    )

    summary = _compute_live_validation_summary(state)
    total = summary.get("total", 0)

    if as_json:
        click.echo(json.dumps(summary, ensure_ascii=False, indent=2))
    else:
        click.echo(f"Состояние: {state_path}")
        click.echo(f"Профиль: {state.get('profile_id', 'gost-7.32-2017')}")
        click.echo("")
        if total == 0:
            click.echo("✓ Нарушений нормоконтроля не найдено.")
        else:
            by_sev = summary.get("by_severity", {})
            click.echo(f"Всего нарушений: {total}")
            click.echo(f"  Ошибок: {by_sev.get('error', 0)}")
            click.echo(f"  Предупреждений: {by_sev.get('warning', 0)}")
            click.echo(f"  Информационных: {by_sev.get('info', 0)}")
            top = summary.get("top_codes", [])
            if top:
                click.echo("")
                click.echo("Топ-5 нарушений:")
                for entry in top:
                    click.echo(f"  {entry['code']}: {entry['count']}")

    # Exit code = 1 если есть error-уровневые нарушения.
    by_sev = summary.get("by_severity", {})
    if by_sev.get("error", 0) > 0:
        sys.exit(1)


@main.command("apply-fixes")
@click.argument("state_path", type=click.Path(exists=True, path_type=Path))
@click.option(
    "-o",
    "--output",
    type=click.Path(path_type=Path),
    required=True,
    help="Куда сохранить исправленный state.",
)
@click.option(
    "--only",
    type=str,
    default=None,
    help="Применить только указанные фиксеры (через запятую, например 'T.08,T.09').",
)
def apply_fixes_cmd(
    state_path: Path, output: Path, only: str | None
) -> None:
    """Применить автофиксы к state и сохранить результат.

    Цикл: state → собрать docx → парсить → fixer.fix() → новый state.
    Применяются все доступные фиксеры из gostforge.fixer (или только
    указанные через --only).

    Пример::

        gostforge apply-fixes draft.json -o fixed.json
        gostforge diff-state draft.json fixed.json    ## посмотреть изменения

    Используется fixer-engine — тот же, что и в кнопке UI
    «Применить автофиксы», но без перезаписи session-state.
    """
    import tempfile  # noqa: PLC0415

    from gostforge.fixer.engine import fix as run_fix  # noqa: PLC0415
    from gostforge.web.builder_editor import (  # noqa: PLC0415
        _build_document_from_state,
        document_to_state,
    )

    try:
        state = json.loads(state_path.read_text(encoding="utf-8"))
    except json.JSONDecodeError as exc:
        click.echo(f"Не удалось прочитать JSON: {exc}", err=True)
        sys.exit(2)
    if not isinstance(state, dict) or "sections" not in state:
        click.echo("В state.json должен быть ключ 'sections'.", err=True)
        sys.exit(2)

    codes = None
    if only:
        codes = [c.strip() for c in only.split(",") if c.strip()]

    try:
        data = _build_document_from_state(state)
    except (FileNotFoundError, ValueError) as exc:
        click.echo(f"Сборка docx упала: {exc}", err=True)
        sys.exit(3)

    with tempfile.NamedTemporaryFile(suffix=".docx", delete=False) as tmp:
        tmp.write(data)
        tmp_path = Path(tmp.name)

    document = parse_docx(tmp_path)
    profile_id = state.get("profile_id", "gost-7.32-2017")
    profile = load_profile(profile_id)
    applied = run_fix(document, profile, codes=codes)
    new_state = document_to_state(document)
    new_state["profile_id"] = profile_id

    output.write_text(
        json.dumps(new_state, ensure_ascii=False, indent=2),
        encoding="utf-8",
    )
    click.echo(f"Применено фиксов: {len(applied)}. Сохранён {output}.")
    if applied:
        # Группируем по коду для компактности.
        from collections import Counter as _Counter  # noqa: PLC0415

        by_code = _Counter(a.fixer_code for a in applied)
        for code, n in sorted(by_code.items()):
            click.echo(f"  {code}: {n}")


@main.command("diff-state")
@click.argument("state_a", type=click.Path(exists=True, path_type=Path))
@click.argument("state_b", type=click.Path(exists=True, path_type=Path))
@click.option(
    "--mode",
    type=click.Choice(["summary", "unified"]),
    default="summary",
    help="Формат вывода: summary — счётчики изменений, unified — построчный diff.",
)
def diff_state_cmd(state_a: Path, state_b: Path, mode: str) -> None:
    """Сравнить два state-файла конструктора.

    Полезно для:
    * сравнения версии до и после редактирования рецензентом;
    * аудита автофикса (что именно поменялось);
    * code-review JSON-state в pull request.

    Режимы:
    * ``summary`` (default) — счётчик добавленных/удалённых/изменённых
      разделов, параграфов, других элементов. Компактный вывод.
    * ``unified`` — построчный diff JSON-представлений с тем же
      форматом, что у ``diff -u``.
    """
    try:
        a = json.loads(state_a.read_text(encoding="utf-8"))
        b = json.loads(state_b.read_text(encoding="utf-8"))
    except json.JSONDecodeError as exc:
        click.echo(f"Не удалось прочитать JSON: {exc}", err=True)
        sys.exit(2)

    if mode == "unified":
        import difflib  # noqa: PLC0415

        a_text = json.dumps(a, ensure_ascii=False, indent=2, sort_keys=True)
        b_text = json.dumps(b, ensure_ascii=False, indent=2, sort_keys=True)
        diff = difflib.unified_diff(
            a_text.splitlines(keepends=True),
            b_text.splitlines(keepends=True),
            fromfile=str(state_a),
            tofile=str(state_b),
        )
        for line in diff:
            click.echo(line, nl=False)
        return

    # summary
    report = _state_diff_summary(a, b)
    if not any(report.values()):
        click.echo("Без изменений.")
        return
    click.echo(f"Сравнение: {state_a} ↔ {state_b}\n")
    if report["title_changed"]:
        click.echo(f"  Заголовок изменён: «{a.get('title')}» → «{b.get('title')}»")
    if report["sections_added"]:
        click.echo(f"  Добавлено разделов: {len(report['sections_added'])}")
        for h in report["sections_added"]:
            click.echo(f"    + {h}")
    if report["sections_removed"]:
        click.echo(f"  Удалено разделов: {len(report['sections_removed'])}")
        for h in report["sections_removed"]:
            click.echo(f"    - {h}")
    if report["sections_modified"]:
        click.echo(f"  Изменено разделов: {len(report['sections_modified'])}")
        for entry in report["sections_modified"]:
            click.echo(f"    ~ {entry['heading']}: {entry['summary']}")


def _state_diff_summary(
    a: dict[str, Any], b: dict[str, Any]
) -> dict[str, Any]:
    """Сравнить два state-словаря, вернуть структуру изменений.

    Маппит разделы по индексу и по заголовку: если в обоих state
    есть раздел с heading=H, считается, что это один и тот же
    раздел (даже если он переместился).
    """
    out: dict[str, Any] = {
        "title_changed": a.get("title") != b.get("title"),
        "sections_added": [],
        "sections_removed": [],
        "sections_modified": [],
    }

    a_sections = a.get("sections") or []
    b_sections = b.get("sections") or []
    a_by_heading = {
        (s.get("heading") or f"#{i}"): s for i, s in enumerate(a_sections)
    }
    b_by_heading = {
        (s.get("heading") or f"#{i}"): s for i, s in enumerate(b_sections)
    }

    a_keys = set(a_by_heading)
    b_keys = set(b_by_heading)
    out["sections_added"] = sorted(b_keys - a_keys)
    out["sections_removed"] = sorted(a_keys - b_keys)
    for key in sorted(a_keys & b_keys):
        summary = _compare_sections(a_by_heading[key], b_by_heading[key])
        if summary:
            out["sections_modified"].append({"heading": key, "summary": summary})
    return out


def _compare_sections(a: dict[str, Any], b: dict[str, Any]) -> str:
    """Вернуть короткое описание изменений двух разделов или '' если идентичны."""
    parts = []
    a_blocks = a.get("blocks") or []
    b_blocks = b.get("blocks") or []
    if len(a_blocks) != len(b_blocks):
        parts.append(f"блоков {len(a_blocks)}→{len(b_blocks)}")
    elif a_blocks != b_blocks:
        parts.append("блоки изменены")
    a_subs = a.get("subsections") or []
    b_subs = b.get("subsections") or []
    if len(a_subs) != len(b_subs):
        parts.append(f"подразделов {len(a_subs)}→{len(b_subs)}")
    a_disabled = sorted(a.get("disabled_checks") or [])
    b_disabled = sorted(b.get("disabled_checks") or [])
    if a_disabled != b_disabled:
        parts.append(f"disabled_checks {a_disabled}→{b_disabled}")
    a_refs = a.get("references") or []
    b_refs = b.get("references") or []
    if len(a_refs) != len(b_refs):
        parts.append(f"источников {len(a_refs)}→{len(b_refs)}")
    elif a_refs != b_refs:
        parts.append("источники изменены")
    return ", ".join(parts)


@main.command("import-md")
@click.argument("md_path", type=click.Path(exists=True, path_type=Path))
@click.option(
    "-o",
    "--output",
    type=click.Path(path_type=Path),
    required=True,
    help="Куда сохранить JSON-state.",
)
@click.option(
    "--title", type=str, default=None,
    help="Название работы (по умолчанию — из первого # заголовка).",
)
@click.option(
    "--profile", "profile_id", type=str, default="gost-7.32-2017",
    help="Идентификатор профиля.",
)
def import_md_cmd(
    md_path: Path,
    output: Path,
    title: str | None,
    profile_id: str,
) -> None:
    """Импортировать Markdown в state-конструктора.

    Обратная операция к ``export-md``. Поддерживает разумное
    подмножество GFM:

    * ``#`` → title работы (если задан один раз в начале);
    * ``##`` / ``###`` → разделы / подразделы;
    * абзацы → kind='paragraph';
    * ``- item`` / ``* item`` → unordered list;
    * ``1. item`` (последовательная нумерация) → ordered list;
    * ``$$ latex $$`` → formula;
    * GFM-таблицы (``| ... |``) → table.

    Не поддерживается: inline-форматирование (`**bold**`, ``*italic*``)
    — текст идёт как plain. Сложные HTML-вставки игнорируются.
    """
    text = md_path.read_text(encoding="utf-8")
    state = _markdown_to_state(text, profile_id=profile_id, title=title)
    output.write_text(
        json.dumps(state, ensure_ascii=False, indent=2),
        encoding="utf-8",
    )
    n = len(state.get("sections", []))
    click.echo(
        f"Импортирован {output} ({n} разделов). Используйте `gostforge ui` "
        f"для редактирования или `generate` для сборки .docx."
    )


def _parse_md_inline(text: str) -> list[dict[str, Any]]:
    """Разобрать inline Markdown в runs (text с bold/italic).

    Поддерживает:
    * '***x***' → bold+italic;
    * '**x**' → bold;
    * '*x*' / '_x_' → italic.

    Простой scanner — для round-trip с export-md достаточно.
    Не поддерживает: code spans (`x`), strikethrough (~~x~~),
    ссылки [text](url) — текст сохраняется как есть.
    """
    import re  # noqa: PLC0415

    # Регекс для трёх вариантов в порядке убывания специфичности:
    # *** ... *** | ** ... ** | * ... * | _ ... _
    pattern = re.compile(
        r"(\*\*\*([^*]+)\*\*\*"
        r"|\*\*([^*]+)\*\*"
        r"|\*([^*]+)\*"
        r"|_([^_]+)_)"
    )
    runs: list[dict[str, Any]] = []
    last_end = 0
    for m in pattern.finditer(text):
        if m.start() > last_end:
            runs.append({"kind": "text", "text": text[last_end : m.start()]})
        if m.group(2) is not None:
            runs.append(
                {"kind": "text", "text": m.group(2), "bold": True, "italic": True}
            )
        elif m.group(3) is not None:
            runs.append({"kind": "text", "text": m.group(3), "bold": True})
        elif m.group(4) is not None:
            runs.append({"kind": "text", "text": m.group(4), "italic": True})
        elif m.group(5) is not None:
            runs.append({"kind": "text", "text": m.group(5), "italic": True})
        last_end = m.end()
    if last_end < len(text):
        runs.append({"kind": "text", "text": text[last_end:]})
    return runs


def _markdown_to_state(
    text: str,
    *,
    profile_id: str = "gost-7.32-2017",
    title: str | None = None,
) -> dict[str, Any]:
    """Распарсить Markdown в state-словарь.

    Простой scanner — не полноценный GFM-парсер. Идём по строкам,
    переключаемся между состояниями (in-table, in-formula).
    Достаточно для round-trip с export-md и большинства руковописных
    .md-файлов.
    """
    import re  # noqa: PLC0415

    lines = text.splitlines()
    state: dict[str, Any] = {
        "title": title or "",
        "year": 2026,
        "profile_id": profile_id,
        "sections": [],
    }
    # Стек открытых секций: (depth, dict). depth = уровень # (2..6).
    # Используется чтобы новый ## закрывал предыдущие подразделы.
    section_stack: list[tuple[int, dict[str, Any]]] = []

    def current_blocks() -> list[dict[str, Any]] | None:
        if not section_stack:
            return None
        return section_stack[-1][1].setdefault("blocks", [])

    i = 0
    while i < len(lines):
        line = lines[i]
        stripped = line.strip()

        # H1: title работы (берём только если ещё не задан).
        m = re.match(r"^#\s+(.+?)\s*$", stripped)
        if m and not stripped.startswith("##"):
            if not state["title"]:
                state["title"] = m.group(1).strip()
            i += 1
            continue

        # H2-H6: секции.
        m = re.match(r"^(#{2,6})\s+(.+?)\s*$", stripped)
        if m:
            depth = len(m.group(1))
            heading = m.group(2).strip()
            new_sec: dict[str, Any] = {"heading": heading, "blocks": []}
            # bibliography по эвристике.
            if heading.lower() in {
                "список использованных источников",
                "список литературы",
                "литература",
                "список источников",
                "библиографический список",
            }:
                new_sec["is_bibliography"] = True
                new_sec["references"] = []
            # Закрываем стек до уровня >= depth.
            while section_stack and section_stack[-1][0] >= depth:
                section_stack.pop()
            if section_stack:
                parent = section_stack[-1][1]
                parent.setdefault("subsections", []).append(new_sec)
            else:
                # depth == 2 (топ-level раздел).
                state["sections"].append(new_sec)
            section_stack.append((depth, new_sec))
            i += 1
            continue

        # GFM-таблица: строка из | ... |, следом разделитель |---|, следом данные.
        if stripped.startswith("|") and i + 1 < len(lines) and re.match(
            r"^\|[\s\-:|]+\|\s*$", lines[i + 1].strip()
        ):
            blocks = current_blocks()
            if blocks is None:
                i += 1
                continue
            headers = [c.strip() for c in stripped.strip("|").split("|")]
            i += 2  # skip header + separator
            rows: list[list[str]] = []
            while i < len(lines) and lines[i].strip().startswith("|"):
                row_cells = [c.strip() for c in lines[i].strip().strip("|").split("|")]
                rows.append(row_cells)
                i += 1
            blocks.append(
                {
                    "kind": "table",
                    "headers": headers,
                    "rows": rows,
                    "caption": "",
                }
            )
            continue

        # Formula: $$ latex $$.
        m = re.match(r"^\$\$\s*(.+?)\s*\$\$\s*$", stripped)
        if m:
            blocks = current_blocks()
            if blocks is not None:
                blocks.append({"kind": "formula", "latex": m.group(1)})
            i += 1
            continue

        # List items: '- text' или '* text' (unordered), '1. text' (ordered).
        m_ul = re.match(r"^[-*]\s+(.+?)\s*$", stripped)
        m_ol = re.match(r"^\d+\.\s+(.+?)\s*$", stripped)
        if m_ul or m_ol:
            ordered = m_ol is not None
            items: list[str] = []
            while i < len(lines):
                s = lines[i].strip()
                m2u = re.match(r"^[-*]\s+(.+?)\s*$", s)
                m2o = re.match(r"^\d+\.\s+(.+?)\s*$", s)
                if ordered and m2o:
                    items.append(m2o.group(1))
                    i += 1
                elif not ordered and m2u:
                    items.append(m2u.group(1))
                    i += 1
                else:
                    break
            # bibliography-секция: items → references.
            if (
                section_stack
                and section_stack[-1][1].get("is_bibliography")
                and ordered
            ):
                section_stack[-1][1].setdefault("references", []).extend(items)
            else:
                blocks = current_blocks()
                if blocks is not None:
                    blocks.append(
                        {"kind": "list", "ordered": ordered, "items": items}
                    )
            continue

        # Figure: ![caption](path).
        m = re.match(r"^!\[([^\]]*)\]\(([^\)]+)\)\s*$", stripped)
        if m:
            blocks = current_blocks()
            if blocks is not None:
                blocks.append(
                    {
                        "kind": "figure",
                        "image_path": m.group(2),
                        "caption": m.group(1),
                    }
                )
            i += 1
            continue

        # Обычный параграф (или пустая строка).
        if stripped:
            blocks = current_blocks()
            if blocks is not None:
                # Если параграф содержит rich-markdown (** / *), разбираем
                # в runs. Иначе кладём как простой text.
                runs = _parse_md_inline(stripped)
                if runs and any(
                    r.get("bold") or r.get("italic") for r in runs
                ):
                    blocks.append({"kind": "paragraph", "runs": runs})
                else:
                    blocks.append({"kind": "paragraph", "text": stripped})
        i += 1

    return state


@main.command("import-docx")
@click.argument("path", type=click.Path(exists=True, path_type=Path))
@click.option(
    "-o",
    "--output",
    type=click.Path(path_type=Path),
    required=True,
    help="Куда сохранить JSON-state для конструктора.",
)
def import_docx_cmd(path: Path, output: Path) -> None:
    """Разложить готовую работу .docx в JSON-state конструктора.

    Это обратная операция к ``gostforge new``: парсит .docx и кладёт
    результат в JSON-формат, который можно загрузить в Streamlit-
    конструктор (sidebar → «Загрузить сохранение (.json)») или
    использовать как промежуточный формат для скриптов.

    Структура JSON: title/author/year/profile_id и sections[] с
    блоками paragraph/table/figure/list/formula. Каждый раздел
    может иметь disabled_checks: list[str] — фича-конструктор
    «не проверять этот раздел нормоконтролем».

    Пример::

        gostforge import-docx work.docx -o draft.json
        gostforge ui
        # В UI: Загрузить сохранение (.json) → draft.json → редактируем

    Ограничения: нумерованные списки текущей реализацией экспорта
    пишутся как обычные параграфы с маркером — при импорте они
    останутся параграфами. Это не теряет содержимое, но требует
    собрать список заново в UI, если он нужен как list-блок.
    """
    from gostforge.web.builder_editor import document_to_state  # noqa: PLC0415

    document = parse_docx(path)
    state = document_to_state(document)
    output.write_text(
        json.dumps(state, ensure_ascii=False, indent=2), encoding="utf-8"
    )
    n_sec = len(state.get("sections", []))
    click.echo(
        f"Разложено {n_sec} разделов в {output}. "
        f"Загрузите его через `gostforge ui` → «Загрузить сохранение (.json)»."
    )


if __name__ == "__main__":
    main()
