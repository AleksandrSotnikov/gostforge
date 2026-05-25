"""Тесты форматов отчётов CLI (Markdown и XLSX)."""

from pathlib import Path

from openpyxl import load_workbook  # type: ignore[import-not-found]

from gostforge.cli import _write_markdown_report, _write_report, _write_xlsx_report
from gostforge.validator import Violation


def _sample() -> dict[str, list[Violation]]:
    """Два файла, один с двумя нарушениями, второй без нарушений."""
    return {
        "/tmp/foo.docx": [
            Violation(
                check_code="F.01",
                severity="error",
                message="Поле top",
                location="page_sections.main",
                suggestion="Установить top = 20",
                details={"expected": "20", "actual": "25"},
            ),
            Violation(
                check_code="T.01",
                severity="warning",
                message="Шрифт Arial",
                location="page_sections.main.paragraph[p-1]",
                suggestion="Использовать Times New Roman",
            ),
        ],
        "/tmp/bar.docx": [],
    }


def test_markdown_report_contains_header_and_table(tmp_path: Path) -> None:
    out = tmp_path / "r.md"
    _write_markdown_report(_sample(), out, "gost-7.32-2017")
    text = out.read_text(encoding="utf-8")
    assert "Отчёт нормоконтроля" in text
    assert "gost-7.32-2017" in text
    assert "foo.docx" in text
    assert "bar.docx" in text
    assert "F.01" in text
    assert "Поле top" in text
    # Категории сгруппированы по префиксу кода
    assert "Категория F" in text and "Категория T" in text


def test_markdown_report_pipe_in_message_is_escaped(tmp_path: Path) -> None:
    violations = {
        "/tmp/x.docx": [
            Violation(
                check_code="T.01",
                severity="error",
                message="Шрифт | внутри сообщения",
                suggestion="Заменить | здесь",
            )
        ]
    }
    out = tmp_path / "r.md"
    _write_markdown_report(violations, out, "gost-7.32-2017")
    text = out.read_text(encoding="utf-8")
    # «|» в тексте должен быть экранирован, чтобы не сломать таблицу
    assert "\\|" in text


def test_xlsx_report_has_summary_and_per_file_sheets(tmp_path: Path) -> None:
    out = tmp_path / "r.xlsx"
    _write_xlsx_report(_sample(), out, "gost-7.32-2017")
    wb = load_workbook(str(out))
    assert "Сводка" in wb.sheetnames
    summary = wb["Сводка"]
    # Заголовочная строка таблицы должна быть в третьей строке листа
    header_row = [c.value for c in summary[3]]
    assert header_row == ["Файл", "Нарушений", "Ошибок", "Предупр.", "Инфо"]
    # Должен быть лист на каждый файл
    assert any("foo" in name for name in wb.sheetnames)
    assert any("bar" in name for name in wb.sheetnames)


def test_xlsx_report_violation_rows(tmp_path: Path) -> None:
    out = tmp_path / "r.xlsx"
    _write_xlsx_report(_sample(), out, "gost-7.32-2017")
    wb = load_workbook(str(out))
    # Найдём лист, относящийся к foo.docx
    foo_sheets = [n for n in wb.sheetnames if "foo" in n]
    assert foo_sheets, "должен быть лист с нарушениями файла foo"
    ws = wb[foo_sheets[0]]
    # 1 строка заголовков + 2 нарушения
    assert ws.max_row == 3
    assert ws["A2"].value == "F.01"
    assert ws["B2"].value == "error"
    assert "Поле top" in ws["C2"].value


def test_write_report_dispatches_by_extension(tmp_path: Path) -> None:
    md = tmp_path / "out.md"
    xlsx = tmp_path / "out.xlsx"
    unknown = tmp_path / "out.txt"

    assert _write_report(_sample(), md, "gost-7.32-2017") == "Markdown"
    assert md.exists()

    assert _write_report(_sample(), xlsx, "gost-7.32-2017") == "Excel"
    assert xlsx.exists()

    # Неизвестное расширение — fallback на Markdown
    assert _write_report(_sample(), unknown, "gost-7.32-2017") == "Markdown"
    assert unknown.exists()


def test_xlsx_report_unique_sheet_names_when_basenames_collide(tmp_path: Path) -> None:
    violations = {
        "/foo/work.docx": [Violation(check_code="F.01", severity="error", message="x")],
        "/bar/work.docx": [Violation(check_code="T.01", severity="error", message="y")],
    }
    out = tmp_path / "r.xlsx"
    _write_xlsx_report(violations, out, "p")
    wb = load_workbook(str(out))
    # Не должно быть пересечений имён листов
    assert len(wb.sheetnames) == len(set(wb.sheetnames))
    # И оба файла должны быть представлены отдельными листами
    assert sum(1 for n in wb.sheetnames if "work" in n) == 2
